import boto3
from botocore.config import Config
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, fonts, PatternFill
from openpyxl.utils.cell import get_column_letter

from django.apps import apps
from django.conf import settings
from django.utils import translation
from django.utils.module_loading import import_string
from django.utils.translation import gettext
from data_ocean import s3bucket

from data_ocean.emails import send_export_url_file_path_message
from users.models import DataOceanUser


class ExportToXlsx:

    @staticmethod
    def export(params, export_dict, model_path, filterset_module, user_id):
        app, model = model_path.split('.')
        queryset = apps.get_model(app, model).objects.all()
        export_filterset = import_string(filterset_module)
        queryset = export_filterset(params,  queryset).qs

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = model
        worksheet.sheet_properties.tabColor = '0033CCCC'
        worksheet.row_dimensions[1].height = 20
        worksheet.freeze_panes = 'A2'
        for col_num, (column_title, column_properties) in enumerate(export_dict.items(), 1):
            row_num = 1
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            worksheet.column_dimensions[get_column_letter(col_num)].width = column_properties[1]
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = fonts.Font(b=True, color='00FFFFFF')
            cell.fill = PatternFill(bgColor='0033CCCC', fill_type="solid")
            for record in queryset:

                row_num += 1
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell_value = getattr(record, column_properties[0])
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.replace(tzinfo=None)
                cell.value = cell_value
        export_file_name = model + '_{0}.xlsx'.format(datetime.now().strftime('%Y-%m-%d_%H-%M-%S'))
        data = BytesIO()
        workbook.save(data)
        data = data.getvalue()

        export_url = s3bucket.save_file(export_file_name, data)
        user = DataOceanUser.objects.get(id=user_id)
        with translation.override(user.language):
            user.notify(
                gettext('Generation of .xlsx file has ended. You may download the file by address %(address)s') % {
                    'address': export_url
                }
            )
            send_export_url_file_path_message(user, export_url)
