from django.core.management import BaseCommand
from openpyxl import load_workbook

from business_register.models.declaration_models import LuxuryCar


class Command(BaseCommand):
    help = 'Saving a list of cars that are subject to transport tax'

    def add_arguments(self, parser):
        pass

    def handle(self, *args, **options):
        FILE_PATH = 'data_ocean/fixtures/'
        FILE_DICT = {
            '2017': 'luxury_cars_2017.xlsx',
            '2018': 'luxury_cars_2018.xlsx',
            '2019': 'luxury_cars_2019.xlsx',
            '2020': 'luxury_cars_2020.xlsx',
            '2021': 'luxury_cars_2021.xlsx'
        }
        FUEL_TYPES = {
            'бензин': LuxuryCar.PETROL,
            'дизель': LuxuryCar.DIESEL,
            'гібрид': LuxuryCar.HYBRID,
            'бензин, електро': LuxuryCar.PETROL_ELECTRIC,
            'гібрид (бензин, електро)': LuxuryCar.PETROL_ELECTRIC,
            'дизель, електро': LuxuryCar.DIESEL_ELECTRIC,
            'бензин (гібрид)': LuxuryCar.HYBRID,
            'електро': LuxuryCar.ELECTRIC,
        }
        for year in FILE_DICT:
            current_year = LuxuryCar.objects.filter(document_year=year).first()
            if current_year:
                continue
            wb = load_workbook(FILE_PATH + FILE_DICT[year])
            ws = wb[wb.sheetnames[0]]
            all_car = []
            for i in range(3, ws.max_row+1):
                car = [cell.value for cell in ws[i]]
                if not car[0]:
                    continue
                brand = car[0].strip().lower()
                model = str(car[1]).strip().lower()
                current_car = [brand, model]
                if current_car in all_car:
                    continue
                doc_year = year
                after_year = int(year) - int(car[2].split(' ')[1])
                volume = float(car[3]) if car[3] and car[3] != 'н/д' and int(car[3]) < 10 else None
                fuel = FUEL_TYPES.get(car[4].lower().strip()) if car[4] else None
                if car[4] and not fuel:
                    self.stdout.write(f'New type of fuel {car[4]}')
                LuxuryCar.objects.create(
                    brand=brand,
                    model=model,
                    after_year=after_year,
                    document_year=doc_year,
                    volume=volume,
                    fuel=fuel,
                )
                all_car.append(current_car)
